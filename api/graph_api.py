import io
import time
import random
from typing import Dict, Optional, List, Tuple

import requests
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import urllib.parse

# =====================================================
# CONFIGURAÇÕES DO AZURE (Secrets do Streamlit)
# =====================================================
CLIENT_ID = st.secrets["AZURE_CLIENT_ID"]
CLIENT_SECRET = st.secrets["AZURE_CLIENT_SECRET"]
TENANT_ID = st.secrets["AZURE_TENANT_ID"]

# =====================================================
# CONFIGURAÇÕES DO SHAREPOINT
# =====================================================
DOMINIO = "osgestora.sharepoint.com"
BIBLIOTECA = "Documentos"
PASTA = "app_refinado_python"
ARQUIVO = "Teste_Refinado - Preenchimento dos gerentes.xlsx"

# =====================================================
# ENDPOINTS DO GRAPH API
# =====================================================
AUTHORITY = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
RESOURCE = "https://graph.microsoft.com/.default"
GRAPH_ROOT = "https://graph.microsoft.com/v1.0"

DEFAULT_TIMEOUT = 25  # segundos

# =====================================================
# Retry / backoff util
# =====================================================

def _backoff_sleep(tentativa: int, base: float = 1.6, jitter: float = 0.8):
    espera = (base ** (tentativa - 1)) + random.uniform(0, jitter)
    time.sleep(min(espera, 10.0))

def _request_with_retry(method: str, url: str, headers: dict = None, **kwargs) -> requests.Response:
    tentativas = kwargs.pop("tentativas", 5)
    for tentativa in range(1, tentativas + 1):
        try:
            resp = requests.request(method, url, headers=headers, timeout=DEFAULT_TIMEOUT, **kwargs)
            resp.raise_for_status()
            return resp
        except requests.exceptions.HTTPError as e:
            status = getattr(getattr(e, "response", None), "status_code", None)
            if status in (423, 429, 502, 503, 504):
                if tentativa == tentativas:
                    raise
                _backoff_sleep(tentativa)
            else:
                raise
        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError):
            if tentativa == tentativas:
                raise
            _backoff_sleep(tentativa)

# =====================================================
# Autenticação e IDs (cacheados)
# =====================================================

@st.cache_resource(show_spinner=False)
def _token_cache() -> dict:
    # armazena {"token": str, "ts": float}
    return {}

def obter_token() -> str:
    cache = _token_cache()
    token = cache.get("token")
    # 58 minutos de margem
    if token and time.time() - cache.get("ts", 0) < 58 * 60:
        return token

    payload = {
        "grant_type": "client_credentials",
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "scope": RESOURCE
    }
    resp = _request_with_retry("POST", AUTHORITY, data=payload)
    token = resp.json()["access_token"]
    cache["token"] = token
    cache["ts"] = time.time()
    return token

@st.cache_resource(show_spinner=False)
def _ids_cache() -> dict:
    # armazena {"site_id": str, "drive_id": str, "item_id": str}
    return {}

def buscar_site_id(token: str) -> str:
    cache = _ids_cache()
    if "site_id" in cache:
        return cache["site_id"]
    url = f"{GRAPH_ROOT}/sites/{DOMINIO}"
    headers = {"Authorization": f"Bearer {token}"}
    resp = _request_with_retry("GET", url, headers=headers)
    cache["site_id"] = resp.json()["id"]
    return cache["site_id"]

def buscar_drive_id(site_id: str, token: str) -> str:
    cache = _ids_cache()
    if "drive_id" in cache:
        return cache["drive_id"]
    url = f"{GRAPH_ROOT}/sites/{site_id}/drives"
    headers = {"Authorization": f"Bearer {token}"}
    resp = _request_with_retry("GET", url, headers=headers)
    for drive in resp.json().get("value", []):
        if drive.get("name") == BIBLIOTECA:
            cache["drive_id"] = drive["id"]
            return cache["drive_id"]
    raise FileNotFoundError(f"Biblioteca '{BIBLIOTECA}' não encontrada.")

def buscar_item_id(site_id: str, drive_id: str, token: str) -> str:
    cache = _ids_cache()
    if "item_id" in cache:
        return cache["item_id"]
    caminho_arquivo = f"{PASTA}/{ARQUIVO}"
    caminho_arquivo_encoded = urllib.parse.quote(caminho_arquivo)
    url = f"{GRAPH_ROOT}/sites/{site_id}/drives/{drive_id}/root:/{caminho_arquivo_encoded}"
    headers = {"Authorization": f"Bearer {token}"}
    resp = _request_with_retry("GET", url, headers=headers)
    rid = resp.json().get("id")
    if not rid:
        raise FileNotFoundError(f"Arquivo '{caminho_arquivo}' não encontrado no SharePoint.")
    cache["item_id"] = rid
    return rid

# =====================================================
# Cache persistente de BYTES + ETag (por sessão)
# =====================================================

@st.cache_resource(show_spinner=False)
def _excel_bytes_store() -> dict:
    """
    Armazena os bytes do arquivo e o ETag para evitar re-download entre páginas:
    {"etag": str | None, "bytes": bytes | None, "last_modified": str | None}
    """
    return {"etag": None, "bytes": None, "last_modified": None}

def _get_item_etag(token: str, site_id: str, drive_id: str, item_id: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Busca metadados mínimos do item (eTag e lastModifiedDateTime) sem baixar o conteúdo.
    """
    url = f"{GRAPH_ROOT}/sites/{site_id}/drives/{drive_id}/items/{item_id}?$select=eTag,lastModifiedDateTime"
    headers = {"Authorization": f"Bearer {token}"}
    resp = _request_with_retry("GET", url, headers=headers)
    js = resp.json()
    return js.get("eTag"), js.get("lastModifiedDateTime")

def _baixar_arquivo_excel_bytes(version_token: int = 0, force: bool = False) -> bytes:
    """
    Retorna os bytes do Excel usando cache por ETag.
    - Se o ETag do SharePoint não mudou, reutiliza os bytes já baixados.
    - Se 'version_token' mudou (após salvamento) ou 'force=True', baixa novamente.
    """
    store = _excel_bytes_store()

    token = obter_token()
    site_id = buscar_site_id(token)
    drive_id = buscar_drive_id(site_id, token)
    item_id = buscar_item_id(site_id, drive_id, token)

    # Se forçar (por salvamento), ignora ETag e baixa tudo
    if force or version_token:
        url = f"{GRAPH_ROOT}/sites/{site_id}/drives/{drive_id}/items/{item_id}/content"
        headers = {"Authorization": f"Bearer {token}"}
        resp = _request_with_retry("GET", url, headers=headers)
        store["bytes"] = resp.content
        # atualiza ETag para refletir o estado atual
        etag, lm = _get_item_etag(token, site_id, drive_id, item_id)
        store["etag"], store["last_modified"] = etag, lm
        return store["bytes"]

    # Consulta rápida do ETag
    etag_remote, lm_remote = _get_item_etag(token, site_id, drive_id, item_id)

    # Se temos bytes e o ETag é o mesmo → reutiliza
    if store.get("bytes") is not None and store.get("etag") == etag_remote:
        return store["bytes"]

    # Caso contrário, baixa bytes e atualiza o store
    url = f"{GRAPH_ROOT}/sites/{site_id}/drives/{drive_id}/items/{item_id}/content"
    headers = {"Authorization": f"Bearer {token}"}
    resp = _request_with_retry("GET", url, headers=headers)
    store["bytes"] = resp.content
    store["etag"] = etag_remote
    store["last_modified"] = lm_remote
    return store["bytes"]

def _bytes_to_excel_file(xls_bytes: bytes) -> pd.ExcelFile:
    return pd.ExcelFile(io.BytesIO(xls_bytes), engine="openpyxl")

# =====================================================
# Leitura de abas (a partir dos bytes cacheados)
# =====================================================

@st.cache_data(ttl=None, show_spinner=False, max_entries=2)
def baixar_arquivo_excel(version_token: int = 0) -> Dict[str, pd.DataFrame]:
    """
    Retorna todas as abas como {nome: DataFrame} a partir de um único download cacheado por ETag.
    """
    content = _baixar_arquivo_excel_bytes(version_token=version_token)
    xls = _bytes_to_excel_file(content)
    sheets = {}
    for name in xls.sheet_names:
        sheets[name] = pd.read_excel(xls, sheet_name=name)
    return sheets

@st.cache_data(ttl=None, show_spinner=False, max_entries=16)
def baixar_aba_excel(nome_aba: str, version_token: int = 0) -> pd.DataFrame:
    """
    Retorna apenas uma aba específica, sem novo download.
    """
    content = _baixar_arquivo_excel_bytes(version_token=version_token)
    xls = _bytes_to_excel_file(content)
    if nome_aba not in xls.sheet_names:
        return pd.DataFrame()
    return pd.read_excel(xls, sheet_name=nome_aba)

# =====================================================
# Escrita (salvar) no Excel
# =====================================================

def _write_df_to_worksheet(ws, df: pd.DataFrame):
    ws.delete_rows(1, ws.max_row if ws.max_row else 1)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

def _abrir_workbook_bytes(version_token: int = 0):
    # Quando version_token>0, forçamos baixar os bytes atuais (sem reutilizar ETag antigo)
    content = _baixar_arquivo_excel_bytes(version_token=version_token, force=bool(version_token))
    return load_workbook(io.BytesIO(content))

def _upload_workbook(wb) -> bool:
    token = obter_token()
    site_id = buscar_site_id(token)
    drive_id = buscar_drive_id(site_id, token)
    item_id = buscar_item_id(site_id, drive_id, token)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    url = f"{GRAPH_ROOT}/sites/{site_id}/drives/{drive_id}/items/{item_id}/content"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }
    _request_with_retry("PUT", url, headers=headers, data=out.read())

    # Atualiza o cache de bytes com o que acabamos de enviar (evita re-download no próximo acesso)
    store = _excel_bytes_store()
    store["bytes"] = out.getvalue()
    etag_new, lm_new = _get_item_etag(token, site_id, drive_id, item_id)
    store["etag"], store["last_modified"] = etag_new, lm_new
    return True

def salvar_arquivo_excel_modificado(sheets_dict: Dict[str, pd.DataFrame], version_token: int = 0) -> bool:
    wb = _abrir_workbook_bytes(version_token=version_token)
    for sheet_name, df in sheets_dict.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
        _write_df_to_worksheet(ws, df)
    return _upload_workbook(wb)

def salvar_apenas_aba(nome_aba: str, df_novo: pd.DataFrame, version_token: int = 0) -> bool:
    wb = _abrir_workbook_bytes(version_token=version_token)
    if nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
    else:
        ws = wb.create_sheet(title=nome_aba)
    _write_df_to_worksheet(ws, df_novo)
    return _upload_workbook(wb)

# =====================================================
# Funções de negócio
# =====================================================

def load_users(version_token: int = 0) -> pd.DataFrame:
    try:
        df_users = baixar_aba_excel("Usuarios", version_token=version_token)
        if df_users is None or df_users.empty:
            return pd.DataFrame(columns=["username", "password_hash", "role", "created_at"])
        return df_users
    except Exception:
        return pd.DataFrame(columns=["username", "password_hash", "role", "created_at"])

def salvar_aba_controle(semana: str, meses_permitidos: Optional[List[str]] = None, version_token: int = 0) -> bool:
    meses_str = ";".join(meses_permitidos) if meses_permitidos else ""
    controle_df = pd.DataFrame({
        "Semana Ativa": [semana],
        "Meses Permitidos": [meses_str]
    })
    return salvar_apenas_aba("Controle", controle_df, version_token=version_token)

def carregar_semana_ativa(version_token: int = 0) -> Optional[dict]:
    try:
        controle_df = baixar_aba_excel("Controle", version_token=version_token)
        if controle_df is None or controle_df.empty:
            return None

        semana = None
        meses_permitidos: List[str] = []
        if "Semana Ativa" in controle_df.columns:
            s = controle_df["Semana Ativa"].dropna()
            if not s.empty:
                semana = str(s.iloc[0])
        if "Meses Permitidos" in controle_df.columns:
            s = controle_df["Meses Permitidos"].dropna()
            if not s.empty:
                meses_str = str(s.iloc[0])
                meses_permitidos = [m.strip() for m in meses_str.split(";") if m.strip()]
        if semana is None:
            return None
        return {"semana": semana, "meses_permitidos": meses_permitidos}
    except Exception:
        return None

def carregar_meses_permitidos(version_token: int = 0) -> List[str]:
    dados = carregar_semana_ativa(version_token=version_token)
    return dados.get("meses_permitidos", []) if dados else []

# =====================================================
# (Opcional) Carregamento stepwise com UI — se quiser usar no primeiro load
# =====================================================

def baixar_aba_excel_stepwise(nome_aba: str, version_token: int = 0, on_update=None) -> pd.DataFrame:
    """
    Igual ao baixar_aba_excel, mas relata etapas via callback 'on_update'.
    Usa o cache por ETag sob o capô, então não baixa novamente se não mudou.
    """
    t0 = time.perf_counter()
    def _say(msg):
        if on_update:
            on_update(msg)

    _say("🔑 Obtendo token…")
    token = obter_token()
    t1 = time.perf_counter(); _say(f"✅ Token em {t1 - t0:.2f}s")

    _say("🧭 Buscando site_id…")
    site_id = buscar_site_id(token)
    t2 = time.perf_counter(); _say(f"✅ site_id em {t2 - t1:.2f}s")

    _say("📁 Buscando drive_id…")
    drive_id = buscar_drive_id(site_id, token)
    t3 = time.perf_counter(); _say(f"✅ drive_id em {t3 - t2:.2f}s")

    _say("📄 Buscando item_id…")
    item_id = buscar_item_id(site_id, drive_id, token)
    t4 = time.perf_counter(); _say(f"✅ item_id em {t4 - t3:.2f}s")

    _say("⬇️ Resolvendo cache de bytes (ETag)…")
    content = _baixar_arquivo_excel_bytes(version_token=version_token)
    t5 = time.perf_counter(); _say(f"✅ Bytes prontos em {t5 - t4:.2f}s")

    _say(f"🧩 Lendo aba '{nome_aba}'…")
    xls = pd.ExcelFile(io.BytesIO(content), engine="openpyxl")
    if nome_aba not in xls.sheet_names:
        _say("⚠️ Aba não encontrada — retornando vazio.")
        return pd.DataFrame()
    df = pd.read_excel(xls, sheet_name=nome_aba)
    t6 = time.perf_counter(); _say(f"✅ Parse em {t6 - t5:.2f}s")
    _say(f"🏁 Concluído em {t6 - t0:.2f}s")
    return df
